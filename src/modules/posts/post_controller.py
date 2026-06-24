from http import HTTPStatus
from io import BytesIO
from fastapi import Depends, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select

from src.common.dependencies import CurrentUser, DbDep
from src.common.pagination import PageQuery, PageResp
from src.common.route import create_router
from src.modules.posts.post_dto import PostBatchCreate, PostCreate, PostResp, PostUpdate
from src.modules.posts.post_model import PostDB
from src.modules.posts.post_service import PostService
from src.modules.users.user_model import UserDB

router = create_router(prefix="/posts", tags=["文章模块"])

svc = PostService()

async def get_own_post(post_id: int, db: DbDep, current_user: CurrentUser):
    db_post = await svc.get_by_id(db, post_id)

    if not db_post:
        raise HTTPException(status_code=HTTPStatus.BAD_REQUEST, detail="文章不存在")

    if db_post.author_id != current_user.id:
        raise HTTPException(status_code=HTTPStatus.FORBIDDEN, detail="该文章不属于您")
    
    return db_post

@router.post("", response_model=PostResp)
async def create_post(payload: PostCreate, current_user: CurrentUser, db: DbDep):
    return await svc.create(db, payload, author_id=current_user.id)

@router.get("/{post_id}", response_model=PostResp)
async def get_post(post_id: int, db: DbDep):
    post = await svc.get_by_id(db, post_id)
    if not post:
        raise HTTPException(status_code=HTTPStatus.BAD_REQUEST, detail="文章不存在")
    return post

@router.get("", response_model=PageResp[PostResp])
async def get_all_posts(db: DbDep, params: PageQuery = Depends(), author_id: int | None = None):
    return await svc.get_all(db, params, author_id=author_id)

@router.put("/{post_id}", response_model=PostResp)
async def update_post(payload: PostUpdate, db: DbDep, db_post: PostDB = Depends(get_own_post)):
    return await svc.update_post(db, payload=payload, db_post=db_post)

@router.delete("/{post_id}")
async def delete_post(db: DbDep, db_post: PostDB = Depends(get_own_post)):    
    return await svc.delete_post(db, db_post=db_post)

@router.post("/batch", response_model=list[PostResp])
async def create_batch_posts(payloads: PostBatchCreate, current_user: CurrentUser, db: DbDep):
    if not payloads.posts:
        raise HTTPException(status_code=HTTPStatus.BAD_REQUEST, detail="文章不能为空")
    return await svc.create_batch(db, payloads=payloads, author_id=current_user.id)

@router.post("/import")
async def import_posts(file: UploadFile, current_user: CurrentUser, db: DbDep):
    content = await file.read()
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    emails = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, content, email = row
        if email and email not in emails:
          emails.append(email)

    stmt = select(UserDB).where(UserDB.email.in_(emails))
    users = (await db.execute(stmt)).scalars().all()

    email_to_user = { user.email: user for user in users }

    to_create = []
    errors = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start = 2):
        title, content, email = row

        if not title:
            errors.append({ "行": idx, "原因": "标题为空" })
            continue

        author = email_to_user.get(email)
        if author is None:
            errors.append({ "行": idx, "原因": f"作者不存在：{email}"})
            continue

        to_create.append(PostDB(title=title, content=content, author_id=author.id))

    db.add_all(to_create)
    await db.commit()
    return {
        "成功": len(to_create),
        "失败": len(errors),
        "失败详情": errors
    }

