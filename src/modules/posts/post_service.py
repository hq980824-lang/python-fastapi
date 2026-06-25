from io import BytesIO
from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from src.common.pagination import PageQuery
from src.modules.posts.post_dto import PostBatchCreate, PostCreate, PostUpdate
from src.modules.posts.post_model import PostDB
from src.modules.users.user_model import UserDB


class PostService:
    async def create(self, db: AsyncSession, payload: PostCreate, author_id: int):
        new_post = PostDB(**payload.model_dump(), author_id=author_id)
        db.add(new_post)
        await db.commit()
        await db.refresh(new_post)
        return new_post

    async def get_by_id(self, db: AsyncSession, post_id: int):
        stmt = (
            select(PostDB)
            .where(PostDB.id == post_id)
            .options(selectinload(PostDB.author))
        )
        result = await db.execute(stmt)
        return result.scalars().first()

    async def get_all(
        self, db: AsyncSession, params: PageQuery, author_id: int | None = None
    ):
        total_stmt = select(func.count(PostDB.id))

        if author_id is not None:
            total_stmt = total_stmt.where(PostDB.author_id == author_id)

        total = await db.scalar(total_stmt)

        offset = (params.page - 1) * params.size
        stmt = (
            select(PostDB)
            .options(selectinload(PostDB.author))
            .offset(offset)
            .limit(params.size)
        )

        if author_id is not None:
            stmt = stmt.where(PostDB.author_id == author_id)

        result = await db.execute(stmt)
        records = result.scalars().all()

        return {
            "records": records,
            "total": total,
            "page": params.page,
            "size": params.size,
        }

    async def update_post(self, db: AsyncSession, payload: PostUpdate, db_post: PostDB):
        update_dict = payload.model_dump(exclude_unset=True)

        for key, value in update_dict.items():
            setattr(db_post, key, value)

        await db.commit()
        return await self.get_by_id(db, db_post.id)

    async def delete_post(self, db: AsyncSession, db_post: PostDB):
        await db.delete(db_post)
        await db.commit()
        return True

    async def create_batch(
        self, db: AsyncSession, payloads: PostBatchCreate, author_id: int
    ):
        new_posts = [
            PostDB(**p.model_dump(), author_id=author_id) for p in payloads.posts
        ]
        db.add_all(new_posts)
        await db.commit()

        ids = [p.id for p in new_posts]
        stmt = select(PostDB).where(PostDB.id.in_(ids))
        result = await db.execute(stmt)
        return result.scalars().all()

    async def import_from_excel(self, db: AsyncSession, file: UploadFile):
        if not file.filename.endswith(".xlsx"):
            raise ValueError('只支持 .xlsx 格式的文件')

        content = await file.read()
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active

        # [[title, content, email]]
        data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

        emails = list(dict.fromkeys([row[-1] for row in data if row[-1]]))

        stmt = select(UserDB).where(UserDB.email.in_(emails))
        users = (await db.execute(stmt)).scalars().all()

        email_to_user = {user.email: user for user in users}

        to_create = []
        errors = []

        for idx, row in enumerate(data, start=2):
            if len(row) < 3:
                errors.append({ "行": idx, "原因": "列数不足"})
                continue

            title, content, email = row[0], row[1], row[2]

            if not title:
                errors.append({"行": idx, "原因": "标题为空"})
                continue

            author = email_to_user.get(email)
            if author is None:
                errors.append({"行": idx, "原因": f"作者不存在：{email}"})
                continue

            to_create.append(PostDB(title=title, content=content, author_id=author.id))

        if not to_create and not errors:
            raise ValueError('文件中没有数据行')

        db.add_all(to_create)
        await db.commit()
        return {"success": len(to_create), "failed": len(errors), "errors": errors}
 
    async def get_all_for_export(self, db: AsyncSession):
        stmt = (
          select(PostDB)
          .options(selectinload(PostDB.author))  
          .order_by(PostDB.id)                   
        )
        result = await db.execute(stmt)
        return result.scalars().all()

    
    async def export_to_excel(self, db: AsyncSession):
        db_posts = await self.get_all_for_export(db)
        
        wb = Workbook()
        ws = wb.active
        ws.append(["标题", "内容", "作者邮箱", "状态"])

        for post in db_posts:
            ws.append([post.title, post.content, post.author.email, post.status.value])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer